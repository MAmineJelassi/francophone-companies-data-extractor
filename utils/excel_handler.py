import os
import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Column definitions for the output workbook
OUTPUT_COLUMNS = [
    'Company',
    'Name',
    'Role',
    'Email',
    'Phone',
    'LinkedIn Profile',
    'Extracted At',
]

SAMPLE_COMPANIES = [
    "Orange Côte d'Ivoire",
    "MTN Côte d'Ivoire",
    "Société Générale Côte d'Ivoire",
    "BICICI",
    "Ecobank Côte d'Ivoire",
    "CFAO Motors Côte d'Ivoire",
    "Total Énergies Côte d'Ivoire",
    "Bolloré Transport & Logistics",
    "Nestlé Côte d'Ivoire",
    "Unilever Côte d'Ivoire",
]


class ExcelHandler:
    """Handles all Excel I/O operations for the automation framework."""

    # ------------------------------------------------------------------ #
    #  Reading                                                             #
    # ------------------------------------------------------------------ #

    def read_companies(self, file_path: str) -> list:
        """
        Read company names from column A of an Excel file.

        If the file does not exist a sample file is created automatically.

        Parameters
        ----------
        file_path : str
            Path to the input Excel file.

        Returns
        -------
        list of str
            Company names (empty rows skipped).
        """
        if not os.path.exists(file_path):
            logger.warning(
                "Input file '%s' not found – creating sample file.", file_path
            )
            self._create_sample_file(file_path)

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        companies = []
        header_skipped = False
        for row in ws.iter_rows(min_col=1, max_col=1, values_only=True):
            cell_value = row[0]
            if cell_value is None:
                continue
            value = str(cell_value).strip()
            if not value:
                continue
            if not header_skipped:
                header_skipped = True
                continue  # skip header row
            companies.append(value)

        wb.close()
        logger.info("Read %d companies from '%s'.", len(companies), file_path)
        return companies

    # ------------------------------------------------------------------ #
    #  Writing                                                             #
    # ------------------------------------------------------------------ #

    def write_results(self, file_path: str, results: list) -> None:
        """
        Write extraction results to a formatted Excel file.

        Parameters
        ----------
        file_path : str
            Destination path for the output file.
        results : list of dict
            Each dict should contain keys matching OUTPUT_COLUMNS.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"

        self._write_header(ws)

        for row_idx, record in enumerate(results, start=2):
            for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=record.get(col_name, 'N/A'))
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if row_idx % 2 == 0:
                    cell.fill = PatternFill(
                        start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'
                    )

        self._auto_fit_columns(ws)
        ws.freeze_panes = 'A2'

        wb.save(file_path)
        logger.info("Results written to '%s' (%d rows).", file_path, len(results))

    # ------------------------------------------------------------------ #
    #  Internal helpers                                                    #
    # ------------------------------------------------------------------ #

    def _write_header(self, ws) -> None:
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(
            start_color='0070C0', end_color='0070C0', fill_type='solid'
        )
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        for col_idx, col_name in enumerate(OUTPUT_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        ws.row_dimensions[1].height = 25

    def _auto_fit_columns(self, ws) -> None:
        for col_cells in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

    def _create_sample_file(self, file_path: str) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Companies"
        ws['A1'] = 'Company Name'
        ws['A1'].font = Font(bold=True)
        for idx, company in enumerate(SAMPLE_COMPANIES, start=2):
            ws.cell(row=idx, column=1, value=company)
        wb.save(file_path)
        logger.info("Sample input file created at '%s'.", file_path)